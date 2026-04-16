from datetime import datetime, time, date
import re

from django.db import transaction
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from main.models import Employee, Attendance
from .models import FingerprintUpload, FingerprintRow
from .serializers import FingerprintUploadSerializer, FingerprintRowSerializer


# =============================================================================
# Access
# =============================================================================

def _is_payroll_admin(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_staff or user.is_superuser:
        return True
    return user.groups.filter(name__in=["Admin", "Manager"]).exists()


def _gate(request):
    if not _is_payroll_admin(request.user):
        return Response({"error": "Admin or Manager access required."}, status=403)
    return None


# =============================================================================
# Parsing helpers
# =============================================================================

NAME_SPLIT_RE = re.compile(r"^\s*([0-9A-Za-z\-]+)\s*\.\s*(.+?)\s*$")


def parse_raw_name(raw):
    """Split '0035 . SANJEEWA PUSHPAKUMARA' → ('35', 'SANJEEWA PUSHPAKUMARA')."""
    if not raw:
        return "", ""
    m = NAME_SPLIT_RE.match(str(raw))
    if not m:
        return "", str(raw).strip()
    code = m.group(1).lstrip("0") or m.group(1)
    name = m.group(2).strip()
    return code, name


def _to_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def _to_time(v):
    if v is None:
        return None
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    s = str(v).strip()
    if s in ("", "-", "—", "N/A"):
        return None
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    return None


def match_employee(parsed_empcode, parsed_name):
    """Return (Employee|None, match_status)."""
    if not parsed_empcode and not parsed_name:
        return None, "Unmatched"

    # 1. Exact empcode (after stripping leading zeros on both sides)
    candidates = []
    if parsed_empcode:
        stripped = parsed_empcode.lstrip("0") or parsed_empcode
        candidates = list(
            Employee.objects.filter(empcode__iexact=stripped) |
            Employee.objects.filter(empcode__iexact=parsed_empcode)
        )

    if candidates and parsed_name:
        for emp in candidates:
            emp_full = (emp.fullname or "").lower()
            for word in parsed_name.lower().split():
                if len(word) >= 3 and word in emp_full:
                    return emp, "Matched"
        return candidates[0], "Ambiguous"

    if candidates:
        return candidates[0], "Ambiguous"

    # 2. Try unique last-word match on fullname
    if parsed_name:
        last_word = parsed_name.split()[-1]
        if len(last_word) >= 3:
            fuzzy = list(Employee.objects.filter(fullname__icontains=last_word))
            if len(fuzzy) == 1:
                return fuzzy[0], "Ambiguous"

    return None, "Unmatched"


def _combine_local(d, t):
    if not d or not t:
        return None
    naive = datetime.combine(d, t)
    return timezone.make_aware(naive)


# =============================================================================
# Upload + parse + stage
# =============================================================================

class FingerprintUploadsAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        gate = _gate(request)
        if gate: return gate
        qs = FingerprintUpload.objects.all()
        return Response(FingerprintUploadSerializer(qs, many=True).data)

    def post(self, request):
        gate = _gate(request)
        if gate: return gate

        f = request.FILES.get("file")
        if not f:
            return Response({"error": "XLS file is required under 'file'."}, status=400)
        if not f.name.lower().endswith((".xlsx", ".xlsm")):
            return Response({"error": "Only .xlsx files are supported."}, status=400)

        try:
            from openpyxl import load_workbook
        except ImportError:
            return Response({"error": "openpyxl is not installed on the server."}, status=500)

        try:
            wb = load_workbook(f, data_only=True, read_only=True)
            ws = wb.active
        except Exception as e:
            return Response({"error": f"Failed to read workbook: {e}"}, status=400)

        # Header row detection — look for "Department","Name","Date","Shift","Time Period","Check-in at","Check-out at"
        header_idx = None
        col_idx = {}
        expected = {"department", "name", "date", "shift", "time period", "check-in at", "check-out at"}
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = [(c or "").strip().lower() if isinstance(c, str) else c for c in row]
            if any(isinstance(v, str) and v == "department" for v in values):
                header_idx = i
                for j, v in enumerate(values):
                    if isinstance(v, str) and v in expected:
                        col_idx[v] = j
                break
            if i > 50:  # don't scan forever
                break

        if header_idx is None or "name" not in col_idx or "date" not in col_idx:
            return Response(
                {"error": "Could not find header row. Expected columns: Department, Name, Date, Shift, Time Period, Check-in at, Check-out at."},
                status=400,
            )

        wb2 = load_workbook(f, data_only=True, read_only=True)
        ws2 = wb2.active

        staged = []
        min_date = None
        max_date = None

        with transaction.atomic():
            upload = FingerprintUpload.objects.create(
                filename=f.name,
                uploaded_by=request.user,
                status="Staged",
            )

            for i, row in enumerate(ws2.iter_rows(values_only=True), start=1):
                if i <= header_idx:
                    continue
                if row is None or all(c is None or c == "" for c in row):
                    continue

                def get(key):
                    idx = col_idx.get(key)
                    return row[idx] if idx is not None and idx < len(row) else None

                raw_name = get("name")
                if raw_name is None or str(raw_name).strip() == "":
                    continue

                parsed_code, parsed_name = parse_raw_name(raw_name)
                d = _to_date(get("date"))
                ci_t = _to_time(get("check-in at"))
                co_t = _to_time(get("check-out at"))

                emp, status_m = match_employee(parsed_code, parsed_name)

                # Build datetime values
                ci_dt = _combine_local(d, ci_t)
                co_dt = _combine_local(d, co_t)

                conflict = False
                if emp and d:
                    conflict = Attendance.objects.filter(employee=emp, date=d).exists()

                staged.append(FingerprintRow(
                    upload=upload,
                    department=str(get("department") or "").strip(),
                    raw_name=str(raw_name).strip(),
                    parsed_empcode=parsed_code,
                    parsed_name=parsed_name,
                    date=d,
                    shift=str(get("shift") or "").strip(),
                    time_period=str(get("time period") or "").strip(),
                    check_in=ci_dt,
                    check_out=co_dt,
                    matched_employee=emp,
                    match_status=status_m,
                    has_asm_conflict=conflict,
                ))

                if d:
                    if min_date is None or d < min_date: min_date = d
                    if max_date is None or d > max_date: max_date = d

            FingerprintRow.objects.bulk_create(staged)

            upload.period_start = min_date
            upload.period_end = max_date
            upload.total_rows = len(staged)
            upload.matched_rows = sum(1 for r in staged if r.match_status == "Matched")
            upload.ambiguous_rows = sum(1 for r in staged if r.match_status == "Ambiguous")
            upload.unmatched_rows = sum(1 for r in staged if r.match_status == "Unmatched")
            upload.conflict_rows = sum(1 for r in staged if r.has_asm_conflict)
            upload.save()

        return Response(FingerprintUploadSerializer(upload).data, status=201)


class FingerprintUploadDetailAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, pk):
        gate = _gate(request)
        if gate: return gate
        upload = get_object_or_404(FingerprintUpload, pk=pk)
        return Response(FingerprintUploadSerializer(upload).data)

    def delete(self, request, pk):
        gate = _gate(request)
        if gate: return gate
        upload = get_object_or_404(FingerprintUpload, pk=pk)
        # Refuse to delete a committed upload — user must revert first
        if upload.status == "Committed":
            return Response(
                {"error": "Upload is Committed. Revert first to delete the staging data."},
                status=400,
            )
        upload.rows.all().delete()
        upload.delete()
        return Response({"message": "Upload removed."}, status=200)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def upload_rows(request, pk):
    gate = _gate(request)
    if gate: return gate
    upload = get_object_or_404(FingerprintUpload, pk=pk)
    qs = upload.rows.select_related("matched_employee", "matched_employee__primary_outlet")

    status_filter = request.GET.get("status")
    if status_filter and status_filter != "all":
        qs = qs.filter(match_status=status_filter)

    conflict = request.GET.get("conflict")
    if conflict == "true":
        qs = qs.filter(has_asm_conflict=True)

    emp_q = request.GET.get("q")
    if emp_q:
        qs = qs.filter(raw_name__icontains=emp_q)

    return Response(FingerprintRowSerializer(qs, many=True).data)


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def row_update(request, pk):
    gate = _gate(request)
    if gate: return gate
    row = get_object_or_404(FingerprintRow, pk=pk)

    if row.upload.status == "Committed" and not request.data.get("_force"):
        return Response({"error": "Parent upload is Committed. Revert before editing rows."}, status=400)

    data = request.data
    if "matched_employee" in data:
        emp_id = data.get("matched_employee")
        if emp_id:
            emp = get_object_or_404(Employee, pk=int(emp_id))
            row.matched_employee = emp
            row.match_status = "Manual"
            row.has_asm_conflict = Attendance.objects.filter(employee=emp, date=row.date).exists() if row.date else False
        else:
            row.matched_employee = None
            row.match_status = "Unmatched"
            row.has_asm_conflict = False

    def parse_dt(val):
        if val is None or val == "":
            return None
        try:
            dt = datetime.fromisoformat(str(val).replace("Z", "+00:00"))
            return dt if dt.tzinfo else timezone.make_aware(dt)
        except (TypeError, ValueError):
            return None

    if "check_in" in data:
        row.check_in = parse_dt(data.get("check_in"))
    if "check_out" in data:
        row.check_out = parse_dt(data.get("check_out"))
    if "skip_commit" in data:
        row.skip_commit = bool(data.get("skip_commit"))
    if "notes" in data:
        row.notes = str(data.get("notes") or "").strip()

    row.save()

    # Recount parent upload stats (cheap)
    upload = row.upload
    rows = upload.rows.all()
    upload.matched_rows = sum(1 for r in rows if r.match_status == "Matched")
    upload.ambiguous_rows = sum(1 for r in rows if r.match_status == "Ambiguous")
    upload.unmatched_rows = sum(1 for r in rows if r.match_status == "Unmatched")
    upload.conflict_rows = sum(1 for r in rows if r.has_asm_conflict)
    upload.save()

    return Response(FingerprintRowSerializer(row).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def rematch(request, pk):
    """Re-run auto-match for all Unmatched/Ambiguous rows."""
    gate = _gate(request)
    if gate: return gate
    upload = get_object_or_404(FingerprintUpload, pk=pk)
    if upload.status == "Committed":
        return Response({"error": "Upload is Committed. Revert to re-match."}, status=400)

    for row in upload.rows.filter(match_status__in=["Unmatched", "Ambiguous"]):
        emp, status_m = match_employee(row.parsed_empcode, row.parsed_name)
        row.matched_employee = emp
        row.match_status = status_m
        row.has_asm_conflict = (
            bool(emp and row.date and Attendance.objects.filter(employee=emp, date=row.date).exists())
        )
        row.save()

    rows = upload.rows.all()
    upload.matched_rows = sum(1 for r in rows if r.match_status == "Matched")
    upload.ambiguous_rows = sum(1 for r in rows if r.match_status == "Ambiguous")
    upload.unmatched_rows = sum(1 for r in rows if r.match_status == "Unmatched")
    upload.conflict_rows = sum(1 for r in rows if r.has_asm_conflict)
    upload.save()

    return Response(FingerprintUploadSerializer(upload).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def commit(request, pk):
    """Create Attendance rows for all (Matched|Manual) & !skip_commit & !conflict rows
    (or rows where override=true applies). Returns counts."""
    gate = _gate(request)
    if gate: return gate
    upload = get_object_or_404(FingerprintUpload, pk=pk)
    if upload.status == "Committed":
        return Response({"error": "Already committed. Revert first."}, status=400)

    override_conflicts = bool(request.data.get("override_conflicts"))

    created = 0
    skipped_conflict = 0
    skipped_unmatched = 0
    skipped_user = 0

    with transaction.atomic():
        for row in upload.rows.select_related("matched_employee").all():
            if row.skip_commit:
                skipped_user += 1
                continue
            if row.match_status not in ("Matched", "Manual") or not row.matched_employee or not row.date:
                skipped_unmatched += 1
                continue
            if row.has_asm_conflict and not override_conflicts:
                skipped_conflict += 1
                continue
            if not row.check_in:
                # Skip rows without a check-in — attendance requires one
                skipped_unmatched += 1
                continue

            att = Attendance.objects.create(
                employee=row.matched_employee,
                date=row.date,
                check_in_time=row.check_in,
                check_in_lat=0.0,
                check_in_long=0.0,
                check_out_time=row.check_out,
                check_out_lat=0.0 if row.check_out else None,
                check_out_long=0.0 if row.check_out else None,
                status="Present",
                punchin_verification="Verified",
                punchout_verification="Verified" if row.check_out else "Pending",
            )
            row.committed_attendance = att
            row.save(update_fields=["committed_attendance"])
            created += 1

        upload.status = "Committed"
        upload.committed_at = timezone.now()
        upload.committed_by = request.user
        upload.save()

    return Response({
        "created": created,
        "skipped_conflict": skipped_conflict,
        "skipped_unmatched": skipped_unmatched,
        "skipped_user": skipped_user,
        "upload": FingerprintUploadSerializer(upload).data,
    })


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def revert(request, pk):
    """Delete all Attendance rows created from this upload, mark upload as Reverted."""
    gate = _gate(request)
    if gate: return gate
    upload = get_object_or_404(FingerprintUpload, pk=pk)
    if upload.status != "Committed":
        return Response({"error": "Upload is not committed."}, status=400)

    removed = 0
    with transaction.atomic():
        for row in upload.rows.filter(committed_attendance__isnull=False):
            att = row.committed_attendance
            row.committed_attendance = None
            row.save(update_fields=["committed_attendance"])
            if att:
                att.delete()
                removed += 1
        upload.status = "Reverted"
        upload.save()

    return Response({"removed": removed, "upload": FingerprintUploadSerializer(upload).data})