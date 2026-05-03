import logging
from decimal import Decimal
from datetime import datetime, date

from django.db import models, transaction

logger = logging.getLogger(__name__)
from django.utils import timezone
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from main.models import Employee, Agency, Outlet
from .models import (
    AllowanceType, AttendanceBonusTier, WorkSchedule,
    Payroll, PayrollAllowance, PayrollDeduction, APITSlab,
    PayrollAuditLog, PayrollCompanyConfig, EmployeeFinancialProfile,
    EpfZone, Bank, BankBranch, AgencyPayrollProfile, OutletEpfPattern,
)
from .serializers import (
    AllowanceTypeSerializer, AttendanceBonusTierSerializer, WorkScheduleSerializer,
    PayrollSerializer, APITSlabSerializer, PayrollAuditLogSerializer,
    PayrollCompanyConfigSerializer, EmployeeFinancialRowSerializer,
    EpfZoneSerializer, BankSerializer, BankBranchSerializer,
    AgencyPayrollProfileSerializer, OutletEpfPatternSerializer,
)
from . import engine


# ─── Access control ────────────────────────────────────────────────────────

PAYROLL_GROUPS = ("Admin", "Acc")
PAYROLL_FEATURE_CODE = "payroll"


def _is_payroll_user(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_staff or user.is_superuser:
        return True
    q = models.Q()
    for name in PAYROLL_GROUPS:
        q |= models.Q(name__iexact=name)
    return user.groups.filter(q).exists()


def _is_admin_user(user):
    if not user or not user.is_authenticated:
        return False
    if user.is_staff or user.is_superuser:
        return True
    return user.groups.filter(name__iexact="Admin").exists()


def _license_feature_ok(request):
    """Return True if the payroll feature is enabled in the current license.
    When no license is bound (request.license is None), allow — the
    LicenseMiddleware treats that as 'active' and we don't want to block dev.
    """
    license_data = getattr(request, "license", None)
    if license_data is None:
        return True
    features = license_data.get("features", []) or []
    return PAYROLL_FEATURE_CODE in features


def _gate(request):
    if not _license_feature_ok(request):
        return Response(
            {"error": "Payroll feature is not enabled on this license.",
             "feature": PAYROLL_FEATURE_CODE},
            status=403,
        )
    if not _is_payroll_user(request.user):
        return Response({"error": "You don't have access to payroll."}, status=403)
    return None


def _audit(payroll, user, action, details=None):
    """Record a payroll audit-log row. Never raises."""
    try:
        PayrollAuditLog.objects.create(
            payroll=payroll,
            user=user if (user and user.is_authenticated) else None,
            action=action,
            details=details or {},
        )
    except Exception:
        logger.exception("Failed to write payroll audit log")


def _parse_date(raw):
    if isinstance(raw, date):
        return raw
    return datetime.strptime(str(raw), "%Y-%m-%d").date()


# ═══════════════════════════════════════════════════════════════════════════
# Allowance Type CRUD
# ═══════════════════════════════════════════════════════════════════════════

class AllowanceTypeListCreate(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        gate = _gate(request)
        if gate: return gate
        qs = AllowanceType.objects.all()
        if request.GET.get("active_only") in ("1", "true", "True"):
            qs = qs.filter(is_active=True)
        return Response(AllowanceTypeSerializer(qs, many=True).data)

    def post(self, request):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin only."}, status=403)
        ser = AllowanceTypeSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data, status=201)


class AllowanceTypeDetail(APIView):
    permission_classes = [IsAuthenticated]

    def _get(self, pk):
        return get_object_or_404(AllowanceType, pk=pk)

    def get(self, request, pk):
        gate = _gate(request)
        if gate: return gate
        return Response(AllowanceTypeSerializer(self._get(pk)).data)

    def patch(self, request, pk):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin only."}, status=403)
        obj = self._get(pk)
        ser = AllowanceTypeSerializer(obj, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)

    def delete(self, request, pk):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin only."}, status=403)
        obj = self._get(pk)
        # Soft-delete via is_active to preserve voucher history
        obj.is_active = False
        obj.save()
        return Response({"message": "Deactivated."})


# ═══════════════════════════════════════════════════════════════════════════
# Attendance Bonus Tier CRUD
# ═══════════════════════════════════════════════════════════════════════════

class BonusTierListCreate(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        gate = _gate(request)
        if gate: return gate
        qs = AttendanceBonusTier.objects.all()
        return Response(AttendanceBonusTierSerializer(qs, many=True).data)

    def post(self, request):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin only."}, status=403)
        ser = AttendanceBonusTierSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data, status=201)


class BonusTierDetail(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin only."}, status=403)
        obj = get_object_or_404(AttendanceBonusTier, pk=pk)
        ser = AttendanceBonusTierSerializer(obj, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)

    def delete(self, request, pk):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin only."}, status=403)
        obj = get_object_or_404(AttendanceBonusTier, pk=pk)
        obj.delete()
        return Response({"message": "Deleted."})


# ═══════════════════════════════════════════════════════════════════════════
# Work Schedule
# ═══════════════════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def work_schedule_list(request):
    gate = _gate(request)
    if gate: return gate

    outlet_id = request.GET.get("outlet")
    qs = (Employee.objects.filter(is_active=True)
          .select_related("primary_outlet", "work_schedule"))
    if outlet_id:
        try:
            qs = qs.filter(outlets__id=int(outlet_id)).distinct()
        except (TypeError, ValueError):
            return Response({"error": "Invalid outlet"}, status=400)
    qs = qs.order_by("fullname")

    rows = []
    for e in qs:
        try:
            s = e.work_schedule
            rows.append({
                "employee_id": e.employee_id,
                "fullname": e.fullname,
                "empcode": e.empcode,
                "primary_outlet_name": e.primary_outlet.name if e.primary_outlet_id else None,
                "configured": True,
                "mon_hours": float(s.mon_hours), "tue_hours": float(s.tue_hours),
                "wed_hours": float(s.wed_hours), "thu_hours": float(s.thu_hours),
                "fri_hours": float(s.fri_hours), "sat_hours": float(s.sat_hours),
                "sun_hours": float(s.sun_hours),
                "ot_multiplier": float(s.ot_multiplier),
                "holiday_multiplier": float(s.holiday_multiplier),
            })
        except WorkSchedule.DoesNotExist:
            rows.append({
                "employee_id": e.employee_id,
                "fullname": e.fullname,
                "empcode": e.empcode,
                "primary_outlet_name": e.primary_outlet.name if e.primary_outlet_id else None,
                "configured": False,
                "mon_hours": 8, "tue_hours": 8, "wed_hours": 8, "thu_hours": 8,
                "fri_hours": 8, "sat_hours": 6, "sun_hours": 0,
                "ot_multiplier": 1.5, "holiday_multiplier": 2.0,
            })
    return Response(rows)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def work_schedule_upsert(request, employee_id):
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    emp = get_object_or_404(Employee, pk=employee_id)
    schedule, _ = WorkSchedule.objects.get_or_create(employee=emp)
    ser = WorkScheduleSerializer(schedule, data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    ser.save()
    return Response(ser.data)


# ═══════════════════════════════════════════════════════════════════════════
# Payroll preview / CRUD
# ═══════════════════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payroll_preview(request, employee_id):
    """Compute a snapshot without saving."""
    gate = _gate(request)
    if gate: return gate
    try:
        sd = _parse_date(request.GET.get("period_start"))
        ed = _parse_date(request.GET.get("period_end"))
    except (TypeError, ValueError):
        return Response({"error": "period_start and period_end (YYYY-MM-DD) required."}, status=400)
    if ed < sd:
        return Response({"error": "period_end must be on or after period_start."}, status=400)

    emp = get_object_or_404(Employee, pk=employee_id)
    rate = request.GET.get("per_day_rate")
    snap = engine.compute(emp, sd, ed, per_day_rate=rate if rate else None)
    # Auto-suggest bonus
    tier, bonus = engine.match_bonus(snap["attendance_score"])

    return Response({
        "employee": {
            "employee_id": emp.employee_id, "fullname": emp.fullname,
            "empcode": emp.empcode, "basic_salary": float(emp.basic_salary or 0),
        },
        "snapshot": {k: (float(v) if hasattr(v, "quantize") else v) for k, v in snap.items()},
        "suggested_bonus": {
            "tier_id": tier.id if tier else None,
            "tier_label": tier.label if tier else None,
            "amount": float(bonus),
        },
    })


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def payroll_list_create(request):
    gate = _gate(request)
    if gate: return gate

    if request.method == "GET":
        qs = Payroll.objects.select_related("employee", "employee__primary_outlet")
        emp_id = request.GET.get("employee_id")
        if emp_id:
            qs = qs.filter(employee_id=emp_id)
        status_f = request.GET.get("status")
        if status_f and status_f != "all":
            qs = qs.filter(status=status_f)
        month = request.GET.get("month")
        if month:
            try:
                y, m = [int(x) for x in month.split("-")[:2]]
                from calendar import monthrange
                last = monthrange(y, m)[1]
                sd = date(y, m, 1); ed = date(y, m, last)
                qs = qs.filter(period_start__lte=ed, period_end__gte=sd)
            except (ValueError, AttributeError):
                return Response({"error": "Invalid month (YYYY-MM)."}, status=400)
        return Response(PayrollSerializer(qs, many=True).data)

    # POST — create Draft payroll from engine snapshot
    data = request.data
    try:
        emp_id = int(data.get("employee_id"))
        sd = _parse_date(data.get("period_start"))
        ed = _parse_date(data.get("period_end"))
    except (TypeError, ValueError):
        return Response({"error": "employee_id, period_start, period_end required."}, status=400)

    emp = get_object_or_404(Employee, pk=emp_id)
    if Payroll.objects.filter(employee=emp, period_start=sd, period_end=ed).exists():
        return Response({"error": f"Payroll already exists for {sd} .. {ed}."}, status=400)

    rate = data.get("per_day_rate")
    snap = engine.compute(emp, sd, ed, per_day_rate=rate if rate else None)

    p = Payroll.objects.create(
        employee=emp,
        period_start=sd, period_end=ed, status=Payroll.STATUS_DRAFT,
        per_day_rate=snap["per_day_rate"], per_hour_rate=snap["per_hour_rate"],
        ot_multiplier=snap["ot_multiplier"], holiday_multiplier=snap["holiday_multiplier"],
        scheduled_hours=snap["scheduled_hours"], worked_hours=snap["worked_hours"],
        ot_hours=snap["ot_hours"], holiday_hours=snap["holiday_hours"],
        holiday_ot_hours=snap["holiday_ot_hours"],
        days_present=snap["days_present"], days_late=snap["days_late"],
        days_half=snap["days_half"], days_absent=snap["days_absent"],
        days_leave=snap["days_leave"], days_holiday_worked=snap["days_holiday_worked"],
        attendance_score=snap["attendance_score"],
        regular_pay=snap["regular_pay"], ot_pay=snap["ot_pay"],
        holiday_pay=snap["holiday_pay"], leave_pay=snap["leave_pay"],
        basic_for_epf=Decimal(str(emp.basic_salary or 0)),
        epf_employee_pct=Decimal(str(emp.epf_emp_per or 8)),
        epf_company_pct=Decimal(str(emp.epf_com_per or 12)),
        etf_company_pct=Decimal(str(emp.etf_com_per or 3)),
        daily_breakdown=snap["daily_breakdown"],
        generated_by=request.user,
    )

    # Auto-apply attendance bonus if a tier matches
    tier, bonus = engine.match_bonus(snap["attendance_score"])
    if tier and bonus > 0:
        PayrollAllowance.objects.create(
            payroll=p, allowance_type=None,
            label=tier.label, amount=bonus,
        )

    engine.recompute_totals(p)
    _audit(p, request.user, "create", {
        "period_start": str(sd), "period_end": str(ed),
        "gross_pay": str(p.gross_pay), "net_pay": str(p.net_pay),
    })
    return Response(PayrollSerializer(p).data, status=201)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def payroll_detail(request, pk):
    gate = _gate(request)
    if gate: return gate

    p = get_object_or_404(
        Payroll.objects.select_related("employee", "employee__primary_outlet"),
        pk=pk,
    )

    if request.method == "GET":
        return Response(PayrollSerializer(p).data)

    if p.status == Payroll.STATUS_LOCKED:
        return Response({"error": "Locked payroll cannot be edited/deleted. Unlock first."}, status=400)

    if request.method == "DELETE":
        _audit(p, request.user, "delete", {
            "period_start": str(p.period_start), "period_end": str(p.period_end),
            "net_pay": str(p.net_pay),
        })
        p.delete()
        return Response({"message": "Deleted."})

    # PATCH — edit allowances / deductions / notes / rate fields
    data = request.data
    editable = [
        "per_day_rate", "per_hour_rate", "ot_multiplier", "holiday_multiplier",
        "regular_pay", "ot_pay", "holiday_pay", "leave_pay",
        "basic_for_epf", "epf_employee_pct", "epf_company_pct", "etf_company_pct",
        "notes",
    ]
    for f in editable:
        if f in data and data[f] is not None:
            setattr(p, f, Decimal(str(data[f])) if f != "notes" else str(data[f]))

    with transaction.atomic():
        if "allowances" in data and isinstance(data["allowances"], list):
            p.allowances.all().delete()
            for a in data["allowances"]:
                at_id = a.get("allowance_type")
                at = AllowanceType.objects.filter(pk=at_id).first() if at_id else None
                amt = Decimal(str(a.get("amount", 0)))
                # Enforce cap if allowance_type provided
                if at and at.max_cap_amount and Decimal(at.max_cap_amount) > 0:
                    if amt > Decimal(at.max_cap_amount):
                        return Response(
                            {"error": f"Amount for '{at.name}' exceeds max cap ({at.max_cap_amount})."},
                            status=400,
                        )
                PayrollAllowance.objects.create(
                    payroll=p, allowance_type=at,
                    label=str(a.get("label") or (at.name if at else "Allowance")).strip(),
                    amount=amt,
                )
        if "deductions" in data and isinstance(data["deductions"], list):
            p.deductions.all().delete()
            for d in data["deductions"]:
                PayrollDeduction.objects.create(
                    payroll=p,
                    label=str(d.get("label", "")).strip() or "Deduction",
                    amount=Decimal(str(d.get("amount", 0))),
                )
        engine.recompute_totals(p)

    _audit(p, request.user, "edit", {
        "allowance_total": str(p.allowance_total),
        "deduction_total": str(p.deduction_total),
        "gross_pay": str(p.gross_pay), "net_pay": str(p.net_pay),
    })
    return Response(PayrollSerializer(p).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def payroll_lock(request, pk):
    gate = _gate(request)
    if gate: return gate
    p = get_object_or_404(Payroll, pk=pk)
    if p.status == Payroll.STATUS_LOCKED:
        return Response({"error": "Already locked."}, status=400)
    engine.recompute_totals(p)
    p.status = Payroll.STATUS_LOCKED
    p.locked_by = request.user
    p.locked_at = timezone.now()
    p.save()
    _audit(p, request.user, "lock", {"net_pay": str(p.net_pay)})
    return Response(PayrollSerializer(p).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def payroll_unlock(request, pk):
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    p = get_object_or_404(Payroll, pk=pk)
    if p.status != Payroll.STATUS_LOCKED:
        return Response({"error": "Not locked."}, status=400)
    p.status = Payroll.STATUS_DRAFT
    p.locked_by = None
    p.locked_at = None
    p.save()
    _audit(p, request.user, "unlock", {})
    return Response(PayrollSerializer(p).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payroll_audit_log(request, pk=None):
    """Audit-trail list. Admin-only. Optional ?payroll=<id> filter, else returns
    the latest 200 entries across all payrolls. `pk` in URL wins if provided."""
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    qs = PayrollAuditLog.objects.select_related("user", "payroll").order_by("-created_at")
    pid = pk or request.GET.get("payroll")
    if pid:
        try:
            qs = qs.filter(payroll_id=int(pid))
        except (TypeError, ValueError):
            return Response({"error": "Invalid payroll id."}, status=400)
    else:
        qs = qs[:200]
    return Response(PayrollAuditLogSerializer(qs, many=True).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payroll_employee_list(request):
    """Employee hub for payroll — with latest payroll status for the given period."""
    gate = _gate(request)
    if gate: return gate

    qs = Employee.objects.filter(is_active=True).select_related("primary_outlet").order_by("fullname")
    outlet_id = request.GET.get("outlet_id")
    if outlet_id:
        try:
            qs = qs.filter(outlets__id=int(outlet_id)).distinct()
        except (TypeError, ValueError):
            return Response({"error": "Invalid outlet_id"}, status=400)

    period_start = request.GET.get("period_start")
    period_end = request.GET.get("period_end")

    rows = []
    for e in qs:
        latest_qs = Payroll.objects.filter(employee=e).order_by("-period_end")
        if period_start and period_end:
            latest_qs = latest_qs.filter(period_start=period_start, period_end=period_end)
        latest = latest_qs.first()
        rows.append({
            "employee_id": e.employee_id,
            "fullname": e.fullname,
            "empcode": e.empcode,
            "primary_outlet_id": e.primary_outlet_id,
            "primary_outlet_name": e.primary_outlet.name if e.primary_outlet_id else None,
            "basic_salary": float(e.basic_salary) if e.basic_salary is not None else 0,
            "payroll_id": latest.id if latest else None,
            "payroll_status": latest.status if latest else None,
            "payroll_net_pay": float(latest.net_pay) if latest else None,
            "payroll_score": float(latest.attendance_score) if latest else None,
            "payroll_period_start": latest.period_start.isoformat() if latest else None,
            "payroll_period_end": latest.period_end.isoformat() if latest else None,
        })
    return Response(rows)


# ═══════════════════════════════════════════════════════════════════════════
# APIT (PAYE) slab CRUD
# ═══════════════════════════════════════════════════════════════════════════

class APITSlabListCreate(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        gate = _gate(request)
        if gate: return gate
        qs = APITSlab.objects.all().order_by("min_monthly")
        return Response(APITSlabSerializer(qs, many=True).data)

    def post(self, request):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin only."}, status=403)
        ser = APITSlabSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data, status=201)


class APITSlabDetail(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin only."}, status=403)
        obj = get_object_or_404(APITSlab, pk=pk)
        ser = APITSlabSerializer(obj, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)

    def delete(self, request, pk):
        if not _is_admin_user(request.user):
            return Response({"error": "Admin only."}, status=403)
        get_object_or_404(APITSlab, pk=pk).delete()
        return Response({"message": "Deleted."})


# ═══════════════════════════════════════════════════════════════════════════
# Gratuity (SL Payment of Gratuity Act — 1/2 month's basic × service years,
# eligible after 5 years). Uses Employee.epf_cal_date as the service-start date.
# ═══════════════════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def gratuity_report(request):
    gate = _gate(request)
    if gate: return gate

    from datetime import date as _date
    today = _date.today()
    outlet_id = request.GET.get("outlet")

    qs = (Employee.objects.filter(is_active=True)
          .select_related("primary_outlet").order_by("fullname"))
    if outlet_id:
        try:
            qs = qs.filter(outlets__id=int(outlet_id)).distinct()
        except (TypeError, ValueError):
            return Response({"error": "Invalid outlet"}, status=400)

    rows = []
    for e in qs:
        start = e.epf_cal_date
        if not start:
            rows.append({
                "employee_id": e.employee_id, "fullname": e.fullname, "empcode": e.empcode,
                "primary_outlet_name": e.primary_outlet.name if e.primary_outlet_id else None,
                "service_start": None, "service_years": 0,
                "basic_salary": float(e.basic_salary or 0),
                "eligible": False, "gratuity": 0, "note": "No EPF cal date on record.",
            })
            continue
        years = (today - start).days / 365.25
        basic = float(e.basic_salary or 0)
        eligible = years >= 5
        gratuity = round((basic / 2) * years, 2) if eligible else 0
        rows.append({
            "employee_id": e.employee_id, "fullname": e.fullname, "empcode": e.empcode,
            "primary_outlet_name": e.primary_outlet.name if e.primary_outlet_id else None,
            "service_start": start.isoformat(),
            "service_years": round(years, 2),
            "basic_salary": basic,
            "eligible": eligible,
            "gratuity": gratuity,
            "note": "" if eligible else "< 5 years — not yet eligible",
        })
    return Response(rows)


# ═══════════════════════════════════════════════════════════════════════════
# Payslip PDF
# ═══════════════════════════════════════════════════════════════════════════

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payslip_pdf(request, pk):
    gate = _gate(request)
    if gate: return gate

    p = get_object_or_404(
        Payroll.objects.select_related("employee", "employee__primary_outlet")
        .prefetch_related("allowances", "deductions"),
        pk=pk,
    )

    import io
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#1976d2"))
    h2 = ParagraphStyle("h2", parent=styles["Heading3"], fontSize=11, textColor=colors.HexColor("#333333"))
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.grey)

    story = []
    story.append(Paragraph("Payslip", h1))
    story.append(Paragraph(
        f"{p.employee.fullname} — {p.employee.empcode or ''}", h2))
    story.append(Paragraph(
        f"Period: {p.period_start} to {p.period_end} &nbsp;&nbsp;|&nbsp;&nbsp; Status: {p.status}",
        small,
    ))
    if p.employee.primary_outlet_id:
        story.append(Paragraph(f"Outlet: {p.employee.primary_outlet.name}", small))
    story.append(Spacer(1, 6 * mm))

    # Attendance summary
    att_tbl = Table([
        ["Scheduled Hrs", "Worked Hrs", "OT Hrs", "Holiday Hrs", "Att. Score"],
        [str(p.scheduled_hours), str(p.worked_hours), str(p.ot_hours),
         str(p.holiday_hours), f"{p.attendance_score}%"],
    ], colWidths=[32*mm]*5)
    att_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(att_tbl)
    story.append(Spacer(1, 5 * mm))

    def _money(n):
        return f"{float(n):,.2f}"

    # Earnings
    earnings = [
        ["Earnings", "Amount (Rs.)"],
        ["Regular Pay", _money(p.regular_pay)],
        ["OT Pay", _money(p.ot_pay)],
        ["Holiday Pay", _money(p.holiday_pay)],
        ["Leave Pay", _money(p.leave_pay)],
    ]
    for a in p.allowances.all():
        earnings.append([f"  {a.label}", _money(a.amount)])
    earnings.append(["Gross", _money(p.gross_pay)])
    e_tbl = Table(earnings, colWidths=[110*mm, 50*mm])
    e_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e3f2fd")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#bbdefb")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(e_tbl)
    story.append(Spacer(1, 4 * mm))

    # Deductions
    deds = [["Deductions", "Amount (Rs.)"]]
    deds.append([f"EPF (Employee {p.epf_employee_pct}%)", _money(p.epf_employee_deduction)])
    if float(p.tax_amount or 0) > 0:
        deds.append([f"APIT ({p.tax_slab_label or 'Tax'})", _money(p.tax_amount)])
    for d in p.deductions.all():
        deds.append([f"  {d.label}", _money(d.amount)])
    deds.append(["Total Deductions",
                 _money(float(p.epf_employee_deduction) + float(p.tax_amount or 0)
                        + float(p.deduction_total))])
    d_tbl = Table(deds, colWidths=[110*mm, 50*mm])
    d_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ffebee")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#ffcdd2")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(d_tbl)
    story.append(Spacer(1, 5 * mm))

    # Net
    net_tbl = Table([["NET PAY", _money(p.net_pay)]], colWidths=[110*mm, 50*mm])
    net_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#1976d2")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 13),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(net_tbl)
    story.append(Spacer(1, 8 * mm))

    # Company contributions (informational)
    info_tbl = Table([
        ["Company Contributions (not deducted from employee)", ""],
        [f"EPF Company ({p.epf_company_pct}%)", _money(p.epf_company_contribution)],
        [f"ETF Company ({p.etf_company_pct}%)", _money(p.etf_company_contribution)],
    ], colWidths=[110*mm, 50*mm])
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.grey),
    ]))
    story.append(info_tbl)

    story.append(Spacer(1, 10 * mm))
    story.append(Paragraph(
        f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M')}. "
        "This is a system-generated payslip and does not require a signature.",
        small,
    ))

    doc.build(story)
    buf.seek(0)

    filename = f"payslip_{p.employee.empcode or p.employee_id}_{p.period_start}_{p.period_end}.pdf"
    resp = HttpResponse(buf.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp



# ═══════════════════════════════════════════════════════════════════════════
# Company config (singleton)
# ═══════════════════════════════════════════════════════════════════════════

@api_view(["GET", "PATCH"])
@permission_classes([IsAuthenticated])
def company_config(request):
    gate = _gate(request)
    if gate: return gate
    cfg = PayrollCompanyConfig.get_solo()
    if request.method == "GET":
        return Response(PayrollCompanyConfigSerializer(cfg).data)
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    ser = PayrollCompanyConfigSerializer(cfg, data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    ser.save()
    return Response(ser.data)


# ─── Per-agency payroll profiles ────────────────────────────────────────────

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def agency_profile_list(request):
    gate = _gate(request)
    if gate: return gate
    if request.method == "GET":
        # Synthesise rows for agencies that don't yet have a profile so the UI
        # can list "everything" in one call.
        existing = {p.agency_id: p for p in AgencyPayrollProfile.objects.select_related("agency").all()}
        rows = []
        for ag in Agency.objects.all().order_by("name"):
            if ag.id in existing:
                rows.append(AgencyPayrollProfileSerializer(existing[ag.id]).data)
            else:
                rows.append({
                    "id": None, "agency": ag.id, "agency_name": ag.name,
                    "company_name": "", "employer_epf_number": "",
                    "employer_etf_number": "", "epf_zone_code": "",
                    "data_submission_number": 1, "last_epf_export_ym": "",
                    "company_bank_name": "", "company_bank_code": "",
                    "company_bank_branch_code": "", "company_bank_account_no": "",
                    "updated_at": None,
                })
        return Response(rows)
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    agency_id = request.data.get("agency")
    if not agency_id:
        return Response({"error": "agency required"}, status=400)
    profile, _ = AgencyPayrollProfile.objects.get_or_create(agency_id=agency_id)
    ser = AgencyPayrollProfileSerializer(profile, data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    ser.save()
    return Response(ser.data, status=201)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def agency_profile_detail(request, pk):
    gate = _gate(request)
    if gate: return gate
    profile = get_object_or_404(AgencyPayrollProfile, pk=pk)
    if request.method == "GET":
        return Response(AgencyPayrollProfileSerializer(profile).data)
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    if request.method == "DELETE":
        profile.delete()
        return Response(status=204)
    ser = AgencyPayrollProfileSerializer(profile, data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    ser.save()
    return Response(ser.data)


# ─── Lookup directories (EPF zones, banks, branches) ───────────────────────

def _lookup_collection(request, qs, ser_cls):
    gate = _gate(request)
    if gate: return gate
    if request.method == "GET":
        return Response(ser_cls(qs, many=True).data)
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    ser = ser_cls(data=request.data)
    ser.is_valid(raise_exception=True)
    ser.save()
    return Response(ser.data, status=201)


def _lookup_item(request, obj, ser_cls):
    gate = _gate(request)
    if gate: return gate
    if request.method == "GET":
        return Response(ser_cls(obj).data)
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    if request.method == "DELETE":
        obj.delete()
        return Response(status=204)
    ser = ser_cls(obj, data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    ser.save()
    return Response(ser.data)


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def epf_zone_list(request):
    return _lookup_collection(request, EpfZone.objects.all(), EpfZoneSerializer)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def epf_zone_detail(request, pk):
    return _lookup_item(request, get_object_or_404(EpfZone, pk=pk), EpfZoneSerializer)


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def bank_list(request):
    return _lookup_collection(request, Bank.objects.all(), BankSerializer)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def bank_detail(request, pk):
    return _lookup_item(request, get_object_or_404(Bank, pk=pk), BankSerializer)


@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def bank_branch_list(request):
    qs = BankBranch.objects.select_related("bank").all()
    bank_id = request.query_params.get("bank")
    if bank_id:
        qs = qs.filter(bank_id=bank_id)
    return _lookup_collection(request, qs, BankBranchSerializer)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def bank_branch_detail(request, pk):
    return _lookup_item(request, get_object_or_404(BankBranch, pk=pk), BankBranchSerializer)


# ─── Outlet EPF pattern + member-number generator ──────────────────────────

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def outlet_epf_pattern_list(request):
    """List one row per outlet, synthesising defaults for outlets without a
    pattern so the directory UI can render everything in one call."""
    gate = _gate(request)
    if gate: return gate
    if request.method == "GET":
        existing = {p.outlet_id: p for p in OutletEpfPattern.objects.select_related("outlet").all()}
        rows = []
        for o in Outlet.objects.all().order_by("name"):
            if o.id in existing:
                rows.append(OutletEpfPatternSerializer(existing[o.id]).data)
            else:
                rows.append({
                    "id": None, "outlet": o.id, "outlet_name": o.name,
                    "prefix": "", "suffix": "", "padding": 4,
                    "next_seq": 1, "is_active": True, "sample": "0001",
                    "updated_at": None,
                })
        return Response(rows)
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    outlet_id = request.data.get("outlet")
    if not outlet_id:
        return Response({"error": "outlet required"}, status=400)
    pattern, _ = OutletEpfPattern.objects.get_or_create(outlet_id=outlet_id)
    ser = OutletEpfPatternSerializer(pattern, data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    ser.save()
    return Response(ser.data, status=201)


@api_view(["GET", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def outlet_epf_pattern_detail(request, pk):
    gate = _gate(request)
    if gate: return gate
    obj = get_object_or_404(OutletEpfPattern, pk=pk)
    if request.method == "GET":
        return Response(OutletEpfPatternSerializer(obj).data)
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    if request.method == "DELETE":
        obj.delete()
        return Response(status=204)
    ser = OutletEpfPatternSerializer(obj, data=request.data, partial=True)
    ser.is_valid(raise_exception=True)
    ser.save()
    return Response(ser.data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def outlet_epf_generate(request, outlet_id):
    """Bulk-issue EPF member numbers for employees in this outlet who don't
    yet have one. Each number is drawn from the outlet's pattern and the
    counter advances atomically. Returns the list of (employee_id, number).
    Pass ?overwrite=1 to reissue numbers for everyone in scope.
    """
    gate = _gate(request)
    if gate: return gate
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    pattern = OutletEpfPattern.objects.filter(outlet_id=outlet_id).first()
    if not pattern or not pattern.is_active:
        return Response({"error": "No active EPF pattern on this outlet."}, status=400)

    overwrite = request.query_params.get("overwrite") in ("1", "true", "True")
    only_ids = request.data.get("employee_ids") or None
    qs = Employee.objects.filter(is_active=True, primary_outlet_id=outlet_id)
    if only_ids:
        qs = qs.filter(employee_id__in=only_ids)
    if not overwrite:
        qs = qs.filter(models.Q(epf_number__isnull=True) | models.Q(epf_number=""))

    issued = []
    with transaction.atomic():
        for e in qs.select_for_update():
            num = pattern.issue()
            e.epf_number = num
            e.save(update_fields=["epf_number"])
            issued.append({"employee_id": e.employee_id, "epf_number": num})
    return Response({"issued": issued, "count": len(issued)})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def employee_epf_generate(request, employee_id):
    """Issue a single EPF number for one employee using their outlet's pattern."""
    gate = _gate(request)
    if gate: return gate
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    e = get_object_or_404(Employee, pk=employee_id)
    if not e.primary_outlet_id:
        return Response({"error": "Employee has no primary outlet."}, status=400)
    pattern = OutletEpfPattern.objects.filter(outlet_id=e.primary_outlet_id, is_active=True).first()
    if not pattern:
        return Response({"error": "No active EPF pattern on this outlet."}, status=400)
    num = pattern.issue()
    e.epf_number = num
    e.save(update_fields=["epf_number"])
    return Response({"employee_id": e.employee_id, "epf_number": num})


# ═══════════════════════════════════════════════════════════════════════════
# Employee financial-profile grid (Employee + FinancialProfile merged)
# ═══════════════════════════════════════════════════════════════════════════

def _derive_initials_surname(fullname):
    parts = [p for p in (fullname or "").split() if p]
    if not parts:
        return "", ""
    surname = parts[-1]
    initials = " ".join(p[0] for p in parts[:-1])
    return initials, surname


def _row_for_employee(e):
    fp = getattr(e, "financial_profile", None)
    if fp is None:
        init, sur = _derive_initials_surname(e.fullname)
        fp_vals = {
            "surname": sur, "initials": init,
            "epf_member_status": EmployeeFinancialProfile.MEMBER_STATUS_EXISTING,
            "etf_member_no": "", "bank_name": "", "bank_code": "",
            "bank_branch_code": "", "bank_account_no": "",
        }
    else:
        fp_vals = {
            "surname": fp.surname, "initials": fp.initials,
            "epf_member_status": fp.epf_member_status,
            "etf_member_no": fp.etf_member_no,
            "bank_name": fp.bank_name, "bank_code": fp.bank_code,
            "bank_branch_code": fp.bank_branch_code,
            "bank_account_no": fp.bank_account_no,
        }
    return {
        "employee_id": e.employee_id,
        "empcode": e.empcode or "",
        "fullname": e.fullname,
        "idnumber": e.idnumber or "",
        "basic_salary": float(e.basic_salary) if e.basic_salary is not None else 0,
        "epf_number": e.epf_number or "",
        "epf_grade": e.epf_grade or "",
        "epf_cal_date": e.epf_cal_date.isoformat() if e.epf_cal_date else None,
        "epf_emp_per": float(e.epf_emp_per) if e.epf_emp_per is not None else 8.0,
        "epf_com_per": float(e.epf_com_per) if e.epf_com_per is not None else 12.0,
        "etf_com_per": float(e.etf_com_per) if e.etf_com_per is not None else 3.0,
        "primary_outlet_name": e.primary_outlet.name if e.primary_outlet_id else None,
        **fp_vals,
    }


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def financial_profile_list(request):
    """Merged Employee + FinancialProfile rows for the grid screens."""
    gate = _gate(request)
    if gate: return gate
    qs = (Employee.objects.filter(is_active=True)
          .select_related("primary_outlet", "financial_profile")
          .order_by("fullname"))
    outlet_id = request.GET.get("outlet_id")
    if outlet_id:
        try:
            qs = qs.filter(outlets__id=int(outlet_id)).distinct()
        except (TypeError, ValueError):
            return Response({"error": "Invalid outlet_id"}, status=400)
    rows = [_row_for_employee(e) for e in qs]
    return Response(rows)


EMPLOYEE_EDITABLE_FIELDS = {
    "empcode", "fullname", "idnumber", "basic_salary",
    "epf_number", "epf_grade", "epf_cal_date",
    "epf_emp_per", "epf_com_per", "etf_com_per",
}
PROFILE_EDITABLE_FIELDS = {
    "surname", "initials", "epf_member_status", "etf_member_no",
    "bank_name", "bank_code", "bank_branch_code", "bank_account_no",
}


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def financial_profile_bulk_save(request):
    """Payload: { rows: [ { employee_id, ...fields } , ... ] }
    Admin only. Updates Employee and FinancialProfile atomically per row.
    Returns: { updated: N, errors: [ {employee_id, error} ] }
    """
    if not _is_admin_user(request.user):
        return Response({"error": "Admin only."}, status=403)
    payload = request.data if isinstance(request.data, dict) else {}
    rows = payload.get("rows") or []
    if not isinstance(rows, list) or not rows:
        return Response({"error": "rows[] required."}, status=400)

    updated = 0
    errors = []
    with transaction.atomic():
        for raw in rows:
            if not isinstance(raw, dict):
                continue
            eid = raw.get("employee_id")
            try:
                eid = int(eid)
            except (TypeError, ValueError):
                errors.append({"employee_id": eid, "error": "Invalid id"})
                continue
            try:
                e = Employee.objects.select_for_update().get(pk=eid)
            except Employee.DoesNotExist:
                errors.append({"employee_id": eid, "error": "Not found"})
                continue

            emp_dirty = False
            for f in EMPLOYEE_EDITABLE_FIELDS:
                if f in raw:
                    val = raw[f]
                    if val == "":
                        val = None if f in {"idnumber", "epf_number", "epf_grade",
                                            "epf_cal_date", "basic_salary"} else ""
                    setattr(e, f, val)
                    emp_dirty = True
            if emp_dirty:
                e.save()

            fp, _created = EmployeeFinancialProfile.objects.get_or_create(employee=e)
            fp_dirty = False
            for f in PROFILE_EDITABLE_FIELDS:
                if f in raw:
                    val = raw[f] if raw[f] is not None else ""
                    setattr(fp, f, val)
                    fp_dirty = True
            if not fp.surname and not fp.initials:
                fp.initials, fp.surname = _derive_initials_surname(e.fullname)
                fp_dirty = True
            if fp_dirty:
                fp.save()
            updated += 1

    return Response({"updated": updated, "errors": errors})


# ═══════════════════════════════════════════════════════════════════════════
# Exports: EPF / ETF / Bank (xlsx)
# ═══════════════════════════════════════════════════════════════════════════

def _int_or_none(v):
    try:
        return int(v) if v not in (None, "", "all") else None
    except (TypeError, ValueError):
        return None


def _parse_month(month_str):
    y, m = month_str.split("-")[:2]
    y, m = int(y), int(m)
    from calendar import monthrange
    last = monthrange(y, m)[1]
    return date(y, m, 1), date(y, m, last), y, m


def _locked_payrolls_for_month(month_str, outlet_id=None, include_drafts=False):
    """Fetch payrolls for a given YYYY-MM. Defaults to Locked-only so a half-
    finished draft can't leak into an EPF/ETF/bank submission."""
    sd, ed, _, _ = _parse_month(month_str)
    qs = (Payroll.objects
          .filter(period_start=sd, period_end=ed)
          .select_related("employee", "employee__financial_profile",
                          "employee__primary_outlet"))
    if not include_drafts:
        qs = qs.filter(status=Payroll.STATUS_LOCKED)
    if outlet_id:
        qs = qs.filter(employee__outlets__id=outlet_id).distinct()
    return qs, sd, ed


def _truthy(v):
    return str(v or "").strip().lower() in ("1", "true", "yes")


def _xlsx_response(wb, filename):
    import io
    from django.http import HttpResponse
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_epf(request):
    gate = _gate(request)
    if gate: return gate
    month = request.GET.get("month")
    if not month:
        return Response({"error": "month=YYYY-MM is required."}, status=400)
    include_drafts = _truthy(request.GET.get("include_drafts"))
    try:
        payrolls, sd, ed = _locked_payrolls_for_month(
            month,
            outlet_id=_int_or_none(request.GET.get("outlet_id")),
            include_drafts=include_drafts,
        )
    except (ValueError, IndexError):
        return Response({"error": "Invalid month format."}, status=400)

    payrolls = list(payrolls)
    if not payrolls:
        return Response(
            {"error": "No payrolls for this month." + (
                "" if include_drafts else " Only Locked payrolls export by default; add ?include_drafts=1 to include Draft.")},
            status=400,
        )

    cfg = PayrollCompanyConfig.get_solo()
    ym = sd.strftime("%Y%m")
    # If this YM was already exported, reuse the same submission number
    # (re-export / correction). Otherwise advance to the next number.
    if cfg.last_epf_export_ym != ym:
        cfg.data_submission_number = (cfg.data_submission_number or 0) + 1 if cfg.last_epf_export_ym else (cfg.data_submission_number or 1)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EPF"
    ws.append([
        "NIC/Passport number", "Last Name", "Initials", "Member AC number",
        "Employer's Contribution Amount (Rs.)", "Member's Contribution Amount (Rs.)",
        "Total Contribution Amount (Rs.)", "Total Earnings (Rs.)",
        "Member Status", "Zone code", "Employer Number",
        "Contribution Year Month", "Data Submission Number",
        "No.of days worked",
        "Occupation Classification Grade(As per the classification of censes  & statistic Dept) ",
    ])
    for p in payrolls:
        e = p.employee
        fp = getattr(e, "financial_profile", None)
        init, sur = (fp.initials, fp.surname) if fp else _derive_initials_surname(e.fullname)
        days = float(p.days_present) + float(p.days_half) * 0.5 + float(p.days_late)
        ws.append([
            e.idnumber or "",
            sur, init,
            e.epf_number or "",
            float(p.epf_company_contribution or 0),
            float(p.epf_employee_deduction or 0),
            float(p.epf_company_contribution or 0) + float(p.epf_employee_deduction or 0),
            float(p.basic_for_epf or 0),
            (fp.epf_member_status if fp else "E"),
            cfg.epf_zone_code or "A",
            cfg.employer_epf_number or "",
            ym,
            cfg.data_submission_number,
            round(days, 2),
            e.epf_grade or "",
        ])

    cfg.last_epf_export_ym = ym
    cfg.save(update_fields=["data_submission_number", "last_epf_export_ym", "updated_at"])

    return _xlsx_response(wb, f"EPF_{ym}.xlsx")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_etf(request):
    gate = _gate(request)
    if gate: return gate
    month = request.GET.get("month")
    if not month:
        return Response({"error": "month=YYYY-MM is required."}, status=400)
    include_drafts = _truthy(request.GET.get("include_drafts"))
    try:
        payrolls, sd, ed = _locked_payrolls_for_month(
            month,
            outlet_id=_int_or_none(request.GET.get("outlet_id")),
            include_drafts=include_drafts,
        )
    except (ValueError, IndexError):
        return Response({"error": "Invalid month format."}, status=400)
    payrolls = list(payrolls)
    if not payrolls:
        return Response(
            {"error": "No payrolls for this month." + (
                "" if include_drafts else " Only Locked payrolls export by default; add ?include_drafts=1 to include Draft.")},
            status=400,
        )
    ym = sd.strftime("%Y%m")

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ETF"
    ws.append([
        "Member Number", "Initials of the name", "Surename only", "NIC Number",
        "Contribution period From", "Contribution period To",
        "Total Contributions in Cents",
    ])
    for p in payrolls:
        e = p.employee
        fp = getattr(e, "financial_profile", None)
        init, sur = (fp.initials, fp.surname) if fp else _derive_initials_surname(e.fullname)
        etf_no = (fp.etf_member_no if fp and fp.etf_member_no else (e.epf_number or ""))
        cents = int(round(float(p.etf_company_contribution or 0) * 100))
        ws.append([etf_no, init, sur, e.idnumber or "", ym, ym, cents])

    return _xlsx_response(wb, f"ETF_{ym}.xlsx")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_bank(request):
    """People's Bank CIB payroll file (6 columns)."""
    gate = _gate(request)
    if gate: return gate
    month = request.GET.get("month")
    if not month:
        return Response({"error": "month=YYYY-MM is required."}, status=400)
    include_drafts = _truthy(request.GET.get("include_drafts"))
    try:
        payrolls, sd, ed = _locked_payrolls_for_month(
            month,
            outlet_id=_int_or_none(request.GET.get("outlet_id")),
            include_drafts=include_drafts,
        )
    except (ValueError, IndexError):
        return Response({"error": "Invalid month format."}, status=400)
    payrolls = list(payrolls)
    if not payrolls:
        return Response(
            {"error": "No payrolls for this month." + (
                "" if include_drafts else " Only Locked payrolls export by default; add ?include_drafts=1 to include Draft.")},
            status=400,
        )
    ym = sd.strftime("%Y-%m")
    narration_tmpl = f"Salary {ym}"

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Bank"
    ws.append([
        "Beneficiary A/C", "Bank Code", "Branch Code", "Amount",
        "Beneficiary Name", "Narration",
    ])
    missing = []
    for p in payrolls:
        e = p.employee
        fp = getattr(e, "financial_profile", None)
        acc = (fp.bank_account_no if fp else "") or ""
        bcode = (fp.bank_code if fp else "") or ""
        branch = (fp.bank_branch_code if fp else "") or ""
        if not acc or not bcode or not branch:
            missing.append(e.fullname or f"#{e.employee_id}")
        ws.append([
            acc, bcode, branch,
            round(float(p.net_pay or 0), 2),
            e.fullname or "",
            narration_tmpl,
        ])

    filename = f"PeoplesBank_Salary_{sd.strftime('%Y%m')}.xlsx"
    resp = _xlsx_response(wb, filename)
    if missing:
        resp["X-Missing-Bank-Details"] = ", ".join(missing[:20])
    return resp
